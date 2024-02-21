# may need to download the excel file first before running this automation script
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait


def update_excel_date(file_path, searchTerm, columnName, new_value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    Dict = {}

    # identify the "price" column in the excel file
    for price_column in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=price_column).value == columnName:          # columnName -->  "price"
            Dict["column"] = price_column         # capture the "price_column" number as it's where the "price" located

    # identify the target "fruit_name" in the excel file
    for each_row in range(2, sheet.max_row+1):
        for each_column in range(1, sheet.max_column+1):
            if sheet.cell(row=each_row, column=each_column).value == searchTerm:        # searchTerm -->  fruit_name
                Dict["row"] = each_row            # capture the "each_row" number as it's where the "fruit_name" located

    # edit the excel with updated value
    sheet.cell(row=Dict["row"], column=Dict["column"]).value = new_value        # new_value --> "500"
    book.save(file_path)            # save the edited excel file to the same file path


# variable
file_path = "/Users/chunholuk/Downloads/download.xlsx"
fruit_name = "Apple"
new_value = "500"

# set-up
driver = webdriver.Chrome()             # not using "Service()" here
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

# download the data file
driver.find_element(By.CSS_SELECTOR, "#downloadButton").click()


# edit the excel with updated value
update_excel_date(file_path, fruit_name, "price", new_value)


# upload the file
# web_element = driver.find_element(locator)
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")   # need "type='file'" attribute to upload
file_input.send_keys(file_path)         # use the ".send_keys" method to pass the file_path to take the file

# validate the upload
wait = WebDriverWait(driver, 5)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")     # parent-child travel; class --> tag
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)         # * -->  de-serialize the "toast_locator" variable


# make the "priceColumn" variable dynamic by getting the value of the attribute using ".get_attribute()" method
priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
print(priceColumn)
# trace back to locate the element from a target field by parent-child travel
actual_price = driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
print(actual_price)

# validate the outcome
assert new_value == actual_price
