# https://pypi.org/project/openpyxl/
import openpyxl


# load the complete excel file; workbook --> collection of sheets
book = openpyxl.load_workbook("/Users/chunholuk/Downloads/Academic/Python/PythonDemo.xlsx")
sheet = book.active                         # select a sheet which is active out of that workbook
cell = sheet.cell(row=1, column=2)          # get cell
print(cell.value)                           # extract the value present in that cell


# assigning value to a cell to write in the excel sheet
sheet.cell(row=2, column=2).value = "byebye@gmail.com"
print(sheet.cell(row=2, column=2).value)
print(sheet["A3"].value)                        # A3 --> row=3, column=1

# get the maximum number of rows present in the sheet
print(sheet.max_row)
# get the maximum number of columns present in the sheet
print(sheet.max_column)


# capture and store the excel data into a dictionary data type
Dict = {}


# print every value present in the sheet
for each_row in range(1, sheet.max_row+1):      # the "sheet.max_row" is the upper cap, so have to +1 here to include it
    # if sheet.cell(row=each_row, column=1).value == "Testcase2":         # scan for it in 1st column from each row
        for each_column in range(1, sheet.max_column+1):
            # # Dict["key"] = value
            # Dict[sheet.cell(row=1, column=each_column).value] = sheet.cell(row=each_row, column=each_column).value
            print(sheet.cell(row=each_row, column=each_column).value)

print(Dict)
