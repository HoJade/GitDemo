import openpyxl


class HomePageData:

    # dictionary  --> key-value pair
    # when pass the key, then the data of the value will be extracted
    test_HomePage_Data = [
        {"email": "hello@gmail.com",
         "password": "123456",
         "name": "hi",
         "2-way data binding text": "bye",
         "gender": "Female"
         },
        {"email": "test@gmail.com",
         "password": "223456",
         "name": "hey",
         "2-way data binding text": "bruh",
         "gender": "Male"
         }
    ]


    # static method -->  to call the method by class.method(), avoid creating object
    @staticmethod
    def getTestData(test_case_name):        # the "self" parameter is required only if the method is non-static

        # load the complete excel file; workbook --> collection of sheets
        book = openpyxl.load_workbook("/Users/chunholuk/Downloads/PythonSelFramework/PythonDemo.xlsx")
        sheet = book.active  # select a sheet which is active out of that workbook

        # capture and store the excel data into a dictionary data type
        Dict = {}

        # extract the data of "test_case_name"
        for each_row in range(1, sheet.max_row + 1):  # the "sheet.max_row" is the upper cap, so have to +1 here

            if sheet.cell(row=each_row, column=1).value == test_case_name:  # scan for it in 1st column from each row

                for each_column in range(2, sheet.max_column + 1):  # start from column=2, so only fetch the test data

                    # Dict["key"] = value
                    Dict[sheet.cell(row=1, column=each_column).value] = sheet.cell(row=each_row,
                                                                                   column=each_column).value
                    print(sheet.cell(row=1, column=each_column).value)  # key
                    print(sheet.cell(row=each_row, column=each_column).value)  # value
        print(Dict)
        return [Dict]    # return the dictionary data type within a list format, because "test_HomePage_Data" is in list
